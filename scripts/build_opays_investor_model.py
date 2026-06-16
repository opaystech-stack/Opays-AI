from __future__ import annotations

from datetime import date
from pathlib import Path
from math import inf

from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties


ROOT = Path(r"C:\LAPOSTE\Projets\OPAYS TECH")
OUT_XLSX = ROOT / "06_ops" / "opays_investor_model.xlsx"
OUT_MD = ROOT / "06_ops" / "opays_investor_model_summary.md"

START_DATE = date(2026, 6, 1)
HORIZON = 36
VAT_RATE = 0.15
CIT_RATE = 0.27
OPENING_CASH = 250_000
SEED_TARGET = 1_200_000
INITIAL_CAPEX = 60_000


offers = [
    {
        "name": "Audit / Diagnostic",
        "price": 25_000,
        "cogs": 0.15,
        "start": 1.0,
        "growth": 0.06,
        "cap": 3.5,
        "notes": "Offre d'entree pour identifier le ROI et les goulots d'etranglement.",
    },
    {
        "name": "Sprint d'Efficience",
        "price": 75_000,
        "cogs": 0.30,
        "start": 0.30,
        "growth": 0.025,
        "cap": 1.4,
        "notes": "Sprint court d'implementation, generalement 2-4 semaines.",
    },
    {
        "name": "Projet d'Implementation",
        "price": 180_000,
        "cogs": 0.45,
        "start": 0.05,
        "growth": 0.02,
        "cap": 0.85,
        "notes": "Execution a haute valeur sur un flux de travail ou systeme client.",
    },
    {
        "name": "Forge Retainer (Accompagnement)",
        "price": 30_000,
        "cogs": 0.20,
        "start": 0.50,
        "growth": 0.10,
        "cap": 4.0,
        "notes": "Optimisation recurrente et support operationnel continu.",
    },
    {
        "name": "Sovereign Retainer (R&D)",
        "price": 45_000,
        "cogs": 0.50,
        "start": 0.00,
        "growth": 0.04,
        "cap": 1.8,
        "notes": "Travaux de R&D, d'autonomie et de resilience strategique.",
    },
    {
        "name": "Dossier Sovereign",
        "price": 120_000,
        "cogs": 0.50,
        "start": 0.00,
        "growth": 0.02,
        "cap": 0.40,
        "notes": "Recherche strategique ponctuelle, note de synthese ou feuille de route.",
    },
]

opex_rows = [
    ("Remuneration Fondateur / CEO", 30_000, 33_000, 36_000),
    ("Support Operations & Gestion", 10_000, 12_000, 14_000),
    ("Marketing & Contenu", 10_000, 12_000, 16_000),
    ("Outils SaaS & Logiciels", 5_000, 6_000, 7_000),
    ("Deplacements & Livraison client", 6_000, 7_000, 8_000),
    ("Frais Juridiques, Comptables & Conformite", 4_500, 5_000, 6_000),
    ("Admin, Internet & Telecoms", 3_500, 4_000, 4_500),
    ("Reserve Imprevus / Contingence", 4_000, 4_000, 5_000),
    ("Reserve Recrutement / Croissance", 5_000, 15_000, 22_000),
]

gtm_rows = [
    ("Page entreprise LinkedIn", "Notoriete et autorite intellectuelle", 8_000, 2.5, 12, 35, 30, "Canal principal de sensibilisation"),
    ("Medias locaux", "Confiance et legitimite locale", 1_000, 8, 25, 40, 35, "Visibilite trimestrielle, credibilite accrue"),
    ("Recommandations / Partenaires", "Opportunites entrantes a forte confiance", 5, 100, 100, 80, 50, "Taux de closing le plus eleve"),
    ("Prospection directe", "Demarchage cible", 40, 10, 60, 25, 40, "Generation active de pipeline"),
]

risks = [
    ("Demande trop lente", "Moyenne", "Pipeline sous l'objectif pendant 2+ mois", "Maintenir des couts fixes bas, intensifier la prospection, packager une offre d'audit plus rapide", "CEO"),
    ("Derive du perimetre (Scope Creep)", "Moyenne", "Projets derivant vers du travail sur mesure", "Garder 3 offres productisees et appliquer une regle stricte de devis complementaire", "Directeur Livraison"),
    ("Dependance au fondateur", "Elevee", "Toutes les ventes et la livraison dependent d'une seule personne", "Documenter les procedures (SOPs), deleguer la livraison, batir un vivier de partenaires", "CEO"),
    ("Tension sur la tresorerie", "Moyenne", "La tresorerie nette descend sous 2 mois de frais de fonctionnement", "Geler les recrutements non critiques, prioriser la vente d'abonnements", "Finance"),
    ("Retard fiscal / conformite", "Faible", "Declaration TVA ou publications legales en retard", "Suivre une checklist mensuelle et faire valider par un expert-comptable externe", "Ops"),
    ("Confusion sur le positionnement", "Moyenne", "Les prospects nous considerent comme une simple agence", "Marteler notre narratif clair : Forge = execution, Sovereign = autonomie", "Marque"),
    ("Derive de l'implementation", "Elevee", "Les projets d'implementation s'allongent sans garde-fous de ROI", "Definir les criteres de succes en amont et vendre au forfait base sur la valeur", "Directeur Livraison"),
]

qa = [
    ("Qu'est-ce qu'Opays Tech ?", "Un cabinet d'ingenierie de l'efficience : nous simplifions les operations, automatisons les taches repetitives et integrons l'IA uniquement la ou elle apporte une valeur mesurable."),
    ("Pourquoi deux branches ?", "Forge est le moteur de cash : diagnostics, implementation, retainers. Sovereign est le moteur d'autonomie : recherche, resilience, intelligence strategique."),
    ("Pourquoi maintenant ?", "Les entreprises subissent des pressions sur leurs couts, les equipes sont surchargees et l'adoption de l'IA reste fragmentee. Il faut passer de la hype a l'execution concrete."),
    ("Qu'est-ce qui rend ce modele credible ?", "Il est oriente services, conscient des marges, productise et fonde sur des gains operationnels reels plutot que sur des promesses vagues d'innovation."),
    ("Comment gagnez-vous de l'argent ?", "Via des audits, des sprints d'efficience, des projets d'implementation, des abonnements d'accompagnement et des dossiers strategiques premium."),
    ("Quelle est la logique tarifaire ?", "Les offres d'entree (audit) ont peu de friction, les sprints et l'implementation capturent la valeur creee, et les retainers apportent la stabilite."),
    ("Comment evitez-vous le piege des agences de services ?", "En vendant des resultats operationnels, en limitant notre catalogue d'offres et en suivant une methode structuree avec des livrables standardises."),
    ("Quelles sont les marges ?", "Le modele vise des marges brutes elevees grace a des methodes repetables et des actifs reutilisables. Le classeur utilise des hypotheses de couts directes prudentes."),
    ("Pourquoi la marge brute n'est-elle pas de 90% ?", "Parce que nous ne masquons pas les couts reels de livraison. Nous reservons une marge saine pour le temps des specialistes et l'execution terrain."),
    ("Quel est le fit avec le marche local ?", "Le marche demande de la clarte, de la fiabilite et de l'automatisation pratique. Notre offre est lisible, realiste et libellee en ZAR."),
    ("Comment vendez-vous ?", "LinkedIn, presse locale, recommandations, prospection directe, et un site web clair qui convertit l'interet en appels de decouverte ou audits."),
    ("Comment convertissez-vous l'attention en revenus ?", "Le parcours est direct : attention -> audit -> sprint -> projet -> retainer. Cette sequence est modélisee dans l'onglet Pipeline."),
    ("Que faire si la demande est plus faible que prevu ?", "L'entreprise peut fonctionner en mode commando (lean). Le classeur inclut un scenario conservateur et maintient les frais de fonctionnement sous controle."),
    ("Comment protegez-vous la qualite de livraison ?", "En standardisant notre methode, en appliquant des checklists strictes et en interdisant les derives de perimetre non facturees."),
    ("Comment utilisez-vous l'IA en interne ?", "Comme un multiplicateur pour nos propres recherches, notre documentation, la memoire de nos projets et le controle qualite."),
    ("Quel est votre fosse defensif (moat) ?", "Un systeme d'exploitation clair, des methodes repetees, notre encrage local, et la reutilisation d'actifs de livraison."),
    ("Pourquoi les investisseurs devraient-ils y croire ?", "Parce que le modele est comprehensible, sobre en capital et evolutif : des services rentables aujourd'hui, de la propriete intellectuelle et des outils demain."),
    ("La TVA est-elle incluse ?", "Le classeur isole la TVA (15%) en pass-through conforme aux exigences de la SARS, sans impact sur le benefice net."),
    ("Pourquoi pas de taxe sur le chiffre d'affaires (turnover tax) ?", "Parce que les services professionnels (conseil/ingenierie) en sont generalement exclus des que les recettes depassent les seuils de la SARS."),
    ("Quelle est l'hypothese d'impot ?", "Nous appliquons de maniere prudente le taux d'impot sur les societes standard de 27% en Afrique du Sud."),
    ("De quel capital avez-vous besoin ?", "Le classeur montre a la fois une trajectoire autofinancee et une trajectoire acceleree par levee de fonds. Nous recommandons un seed de 1 200 000 ZAR."),
    ("Quelle sera l'utilisation des fonds ?", "Principalement le renforcement commercial, la capacite de livraison, la standardisation et notre reserve de roulement."),
    ("Quel est le plus grand risque ?", "La dependance vis-a-vis du fondateur. Notre remede est la documentation systematique des process et la delegation progressive."),
    ("Qu'est-ce que les clients achetent en premier ?", "L'audit / diagnostic. C'est l'entree la plus naturelle et notre premiere preuve de valeur."),
    ("Comment gerez-vous les dossiers Sovereign ?", "Comme des mandats de recherche et de conseil strategique haut de gamme, et non comme de la sous-traitance standard."),
]


def month_label(idx: int) -> str:
    year = START_DATE.year + (START_DATE.month - 1 + idx) // 12
    month = (START_DATE.month - 1 + idx) % 12 + 1
    return f"{year}-{month:02d}"


def units_for(offer, m):
    return min(offer["cap"], offer["start"] + offer["growth"] * (m - 1))


def phase_value(ph1, ph2, ph3, m):
    if m <= 12:
        return ph1
    if m <= 24:
        return ph2
    return ph3


def compute_base_model(multiplier=1.0, opex_mult=1.0):
    months = list(range(1, HORIZON + 1))
    rev, cogs, gross, opex, ebitda, tax, net, cash = [], [], [], [], [], [], [], []
    revenue_by_offer = {o["name"]: [] for o in offers}
    cogs_by_offer = {o["name"]: [] for o in offers}
    units_by_offer = {o["name"]: [] for o in offers}
    capex = []

    running_cash = OPENING_CASH
    for m in months:
        total_rev = total_cogs = 0
        for o in offers:
            u = units_for(o, m) * multiplier
            r = u * o["price"]
            cc = r * o["cogs"]
            units_by_offer[o["name"]].append(u)
            revenue_by_offer[o["name"]].append(r)
            cogs_by_offer[o["name"]].append(cc)
            total_rev += r
            total_cogs += cc

        gp = total_rev - total_cogs
        fixed = sum(phase_value(r[1], r[2], r[3], m) for r in opex_rows) * opex_mult
        e = gp - fixed
        t = max(e, 0) * CIT_RATE
        n = e - t
        c = INITIAL_CAPEX if m == 1 else (20_000 if m == 13 else (15_000 if m == 25 else 0))
        running_cash = running_cash + e - c  # no external financing in base case
        rev.append(total_rev)
        cogs.append(total_cogs)
        gross.append(gp)
        opex.append(fixed)
        ebitda.append(e)
        tax.append(t)
        net.append(n)
        capex.append(c)
        cash.append(running_cash)
    return {
        "months": months,
        "rev": rev,
        "cogs": cogs,
        "gross": gross,
        "opex": opex,
        "ebitda": ebitda,
        "tax": tax,
        "net": net,
        "cash": cash,
        "units": units_by_offer,
        "revenue_by_offer": revenue_by_offer,
        "cogs_by_offer": cogs_by_offer,
        "capex": capex,
    }


base = compute_base_model()
conservative = compute_base_model(multiplier=0.82, opex_mult=1.05)
aggressive = compute_base_model(multiplier=1.18, opex_mult=1.10)


def total(series):
    return sum(series)


wb = Workbook()
wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
ws0 = wb.active
ws0.title = "Dashboard"
ws_readme = wb.create_sheet("README")
ws_ass = wb.create_sheet("Assumptions")
ws_offers = wb.create_sheet("Offers")
ws_gtm = wb.create_sheet("GTM")
ws_rev = wb.create_sheet("Revenue_Model")
ws_exp = wb.create_sheet("Expenses")
ws_pnl = wb.create_sheet("P&L")
ws_cash = wb.create_sheet("Cashflow")
ws_fund = wb.create_sheet("Funding")
ws_scn = wb.create_sheet("Scenarios")
ws_risk = wb.create_sheet("Risks_QA")
ws_src = wb.create_sheet("Sources")
ws_ops = wb.create_sheet("Operations")
ws_cap = wb.create_sheet("Capacity")
ws_del = wb.create_sheet("Delivery")
ws_cad = wb.create_sheet("Cadence")
ws_sop = wb.create_sheet("SOPs")
ws_team = wb.create_sheet("Team")
ws_pipe = wb.create_sheet("Pipeline")
ws_90 = wb.create_sheet("90_Day_Plan")
ws_pitch = wb.create_sheet("Pitch_Outline")
ws_qa = wb.create_sheet("Investor_QA")


nav_fill = PatternFill("solid", fgColor="0F172A")
nav_text = Font(color="FFFFFF", bold=True)
section_fill = PatternFill("solid", fgColor="E2E8F0")
gold_fill = PatternFill("solid", fgColor="F59E0B")
green_fill = PatternFill("solid", fgColor="DCFCE7")
red_fill = PatternFill("solid", fgColor="FEE2E2")
blue_fill = PatternFill("solid", fgColor="DBEAFE")
thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = nav_fill
            cell.font = nav_text
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def fmt_money(cell):
    cell.number_format = '#,##0.00_);[Red](#,##0.00)'


def fmt_int(cell):
    cell.number_format = '#,##0'


def fmt_pct(cell):
    cell.number_format = '0.0%'


def set_title(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="0F172A")
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="334155")


# README
set_title(ws_readme, "Classeur Investisseur Opays Tech", "Un outil unique regroupant la strategie, le modele financier, la gestion des risques et la preparation operationnelle.")
readme_lines = [
    "Objectif : Structurer le developpement d'Opays Tech en un modele operationnel et financier robuste, sans surevaluation.",
    "Narratif : Forge = moteur de cash, Sovereign = developpement d'actifs et R&D, le tout pilote par un systeme d'exploitation lean.",
    "Devise : Rand sud-africain (ZAR).",
    "Remarque fiscale importante : Le modele applique de maniere prudente les regles fiscales standards et comptabilise la TVA separement.",
    "Utilisation : Consultez d'abord le Tableau de bord, puis les Hypotheses, puis les details financiers et operationnels pour la livraison.",
]
for i, line in enumerate(readme_lines, start=4):
    ws_readme[f"A{i}"] = f"• {line}"
ws_readme["A11"] = "Onglets inclus dans ce classeur :"
for i, sh in enumerate(["Dashboard", "Assumptions", "Offers", "GTM", "Revenue_Model", "Expenses", "P&L", "Cashflow", "Funding", "Scenarios", "Risks_QA", "Sources"], start=12):
    ws_readme[f"A{i}"] = f"- {sh}"
ops_extra_row = 24
ws_readme[f"A{ops_extra_row}"] = "Onglets de gestion operationnelle add-on :"
for i, sh in enumerate(["Operations", "Capacity", "Delivery", "Cadence", "SOPs", "Team", "Pipeline", "90_Day_Plan", "Pitch_Outline", "Investor_QA"], start=ops_extra_row + 1):
    ws_readme[f"A{i}"] = f"- {sh}"
ws_readme.column_dimensions["A"].width = 120


# Assumptions
set_title(ws_ass, "Hypotheses generales", "Entrees modifiables. Ajustez ces valeurs pour adapter ou localiser le modele.")
general = [
    ("Devise de reference", "ZAR"),
    ("Debut des previsions", START_DATE),
    ("Duree des previsions (mois)", HORIZON),
    ("Taux de TVA", VAT_RATE),
    ("Taux d'impot sur les societes", CIT_RATE),
    ("Tresorerie initiale", OPENING_CASH),
    ("Cible de financement recommandee (Seed)", SEED_TARGET),
    ("Investissement initial (Mois 1)", INITIAL_CAPEX),
]
for idx, (label, value) in enumerate(general, start=4):
    ws_ass[f"A{idx}"] = label
    ws_ass[f"B{idx}"] = value
    ws_ass[f"A{idx}"].font = Font(bold=True)
    ws_ass[f"A{idx}"].fill = section_fill
    ws_ass[f"A{idx}"].border = border
    ws_ass[f"B{idx}"].border = border
fmt_pct(ws_ass["B7"]); fmt_pct(ws_ass["B8"])
fmt_money(ws_ass["B9"]); fmt_money(ws_ass["B10"]); fmt_money(ws_ass["B11"])

offer_start_row = 13
ws_ass[f"A{offer_start_row}"] = "Modele des Offres"
ws_ass[f"A{offer_start_row}"].font = Font(bold=True)
ws_ass[f"A{offer_start_row}"].fill = blue_fill
for c, h in enumerate(["Offre", "Tarif (ZAR)", "Cout direct", "Marge brute", "Remarques"], start=1):
    ws_ass.cell(row=offer_start_row + 1, column=c, value=h)
style_header(ws_ass, f"A{offer_start_row+1}:E{offer_start_row+1}")
offer_row_map = {}
for r, o in enumerate(offers, start=offer_start_row + 2):
    offer_row_map[o["name"]] = r
    ws_ass.cell(row=r, column=1, value=o["name"])
    ws_ass.cell(row=r, column=2, value=o["price"])
    ws_ass.cell(row=r, column=3, value=o["cogs"])
    ws_ass.cell(row=r, column=4, value=f"=1-C{r}")
    ws_ass.cell(row=r, column=5, value=o["notes"])
    fmt_money(ws_ass.cell(row=r, column=2))
    fmt_pct(ws_ass.cell(row=r, column=3))
    fmt_pct(ws_ass.cell(row=r, column=4))

driver_start_row = offer_start_row + 10
ws_ass[f"A{driver_start_row}"] = "Facteurs de croissance"
ws_ass[f"A{driver_start_row}"].font = Font(bold=True)
ws_ass[f"A{driver_start_row}"].fill = blue_fill
for c, h in enumerate(["Offre", "Unites initiales", "Croissance mensuelle", "Plafond"], start=1):
    ws_ass.cell(row=driver_start_row + 1, column=c, value=h)
style_header(ws_ass, f"A{driver_start_row+1}:D{driver_start_row+1}")
driver_row_map = {}
for r, o in enumerate(offers, start=driver_start_row + 2):
    driver_row_map[o["name"]] = r
    ws_ass.cell(row=r, column=1, value=o["name"])
    ws_ass.cell(row=r, column=2, value=o["start"])
    ws_ass.cell(row=r, column=3, value=o["growth"])
    ws_ass.cell(row=r, column=4, value=o["cap"])
    fmt_pct(ws_ass.cell(row=r, column=3))

opex_start_row = driver_start_row + 10
ws_ass[f"A{opex_start_row}"] = "Phases d'Opex fixes"
ws_ass[f"A{opex_start_row}"].font = Font(bold=True)
ws_ass[f"A{opex_start_row}"].fill = blue_fill
for c, h in enumerate(["Categorie d'Opex", "Mois 1-12", "Mois 13-24", "Mois 25-36"], start=1):
    ws_ass.cell(row=opex_start_row + 1, column=c, value=h)
style_header(ws_ass, f"A{opex_start_row+1}:D{opex_start_row+1}")
opex_row_map = {}
for r, (label, p1, p2, p3) in enumerate(opex_rows, start=opex_start_row + 2):
    opex_row_map[label] = r
    ws_ass.cell(row=r, column=1, value=label)
    ws_ass.cell(row=r, column=2, value=p1)
    ws_ass.cell(row=r, column=3, value=p2)
    ws_ass.cell(row=r, column=4, value=p3)
    fmt_money(ws_ass.cell(row=r, column=2))
    fmt_money(ws_ass.cell(row=r, column=3))
    fmt_money(ws_ass.cell(row=r, column=4))

gtm_start_row = opex_start_row + len(opex_rows) + 4
ws_ass[f"A{gtm_start_row}"] = "Hypotheses Go-To-Market"
ws_ass[f"A{gtm_start_row}"].font = Font(bold=True)
ws_ass[f"A{gtm_start_row}"].fill = blue_fill
for c, h in enumerate(["Canal", "Objectif", "Activite mensuelle", "Taux Prospect", "Taux Appel", "Taux Closing", "Notes"], start=1):
    ws_ass.cell(row=gtm_start_row + 1, column=c, value=h)
style_header(ws_ass, f"A{gtm_start_row+1}:G{gtm_start_row+1}")
for r, row in enumerate(gtm_rows, start=gtm_start_row + 2):
    for c, val in enumerate(row, start=1):
        ws_ass.cell(row=r, column=c, value=val)
    fmt_pct(ws_ass.cell(row=r, column=3))
    fmt_pct(ws_ass.cell(row=r, column=4))
    fmt_pct(ws_ass.cell(row=r, column=5))
    fmt_pct(ws_ass.cell(row=r, column=6))

fund_row = gtm_start_row + len(gtm_rows) + 4
ws_ass[f"A{fund_row}"] = "Contexte de financement"
ws_ass[f"A{fund_row}"].font = Font(bold=True)
ws_ass[f"A{fund_row}"].fill = blue_fill
funding_entries = [
    ("Objectif minimum (Runway lean)", 300_000),
    ("Financement recommande (Seed)", SEED_TARGET),
    ("Affectation : Ventes & Image de marque", 0.25),
    ("Affectation : Capacite de livraison", 0.30),
    ("Affectation : Standardisation & IP", 0.20),
    ("Affectation : Conformite & Administration", 0.10),
    ("Affectation : Reserve de roulement", 0.15),
]
for i, (label, value) in enumerate(funding_entries, start=fund_row + 1):
    ws_ass[f"A{i}"] = label
    ws_ass[f"B{i}"] = value
    ws_ass[f"A{i}"].font = Font(bold=True)
    ws_ass[f"A{i}"].fill = section_fill
    ws_ass[f"A{i}"].border = border
    ws_ass[f"B{i}"].border = border
    if isinstance(value, float):
        fmt_pct(ws_ass[f"B{i}"])
    else:
        fmt_money(ws_ass[f"B{i}"])

source_row = fund_row + len(funding_entries) + 4
ws_ass[f"A{source_row}"] = "Sources / Notes fiscales"
ws_ass[f"A{source_row}"].font = Font(bold=True)
ws_ass[f"A{source_row}"].fill = blue_fill
source_entries = [
    ("TVA", "SARS : taux de TVA standard de 15%. Enregistrement requis selon seuils SARS.", "https://www.sars.gov.za/types-of-tax/value-added-tax/"),
    ("IS", "SARS : taux standard d'impot sur les societes de 27% pour 2026/27.", "https://www.sars.gov.za/types-of-tax/corporate-income-tax/"),
    ("Taxe forfaitaire micro-entreprises", "SARS : les services professionnels comme le conseil et l'ingenierie en sont generalement exclus.", "https://www.sars.gov.za/faq/faq-what-are-the-requirements-for-a-micro-business-to-qualify-for-turnover-tax/"),
]
for i, (topic, note, url) in enumerate(source_entries, start=source_row + 1):
    ws_ass[f"A{i}"] = topic
    ws_ass[f"B{i}"] = note
    ws_ass[f"C{i}"] = url
    ws_ass[f"A{i}"].font = Font(bold=True)
    ws_ass[f"A{i}"].alignment = Alignment(wrap_text=True)
    ws_ass[f"B{i}"].alignment = Alignment(wrap_text=True)
    ws_ass[f"C{i}"].alignment = Alignment(wrap_text=True)

for col, width in {"A": 38, "B": 18, "C": 18, "D": 18, "E": 56, "F": 20, "G": 20}.items():
    ws_ass.column_dimensions[col].width = width


# Offers sheet
set_title(ws_offers, "Architecture des Offres", "Offres productisees avec tarifications, livraison et marges conservatives.")
for c, h in enumerate(["Offre", "Tarif (ZAR)", "Cout direct %", "Marge brute %", "Probleme client traite", "Notes de livraison"], start=1):
    ws_offers.cell(row=4, column=c, value=h)
style_header(ws_offers, "A4:F4")
for r, o in enumerate(offers, start=5):
    ws_offers.cell(row=r, column=1, value=o["name"])
    ws_offers.cell(row=r, column=2, value=o["price"])
    ws_offers.cell(row=r, column=3, value=o["cogs"])
    ws_offers.cell(row=r, column=4, value=f"=1-C{r}")
    ws_offers.cell(row=r, column=5, value={
        "Audit / Diagnostic": "Identifier les goulots d'etranglement, cartographier les frictions et reveler le ROI.",
        "Sprint d'Efficience": "Resoudre un probleme precis de flux de travail rapidement et de maniere visible.",
        "Projet d'Implementation": "Transformer durablement un flux de travail ou integrer un systeme sur mesure.",
        "Forge Retainer (Accompagnement)": "Fournir un support et une optimisation continue des systemes installes.",
        "Sovereign Retainer (R&D)": "Soutenir la resilience strategique et l'autonomie technologique locale.",
        "Dossier Sovereign": "Fournir une recherche strategique approfondie ou une note de synthese prete pour le Board.",
    }[o["name"]])
    ws_offers.cell(row=r, column=6, value=o["notes"])
    fmt_money(ws_offers.cell(row=r, column=2))
    fmt_pct(ws_offers.cell(row=r, column=3))
    fmt_pct(ws_offers.cell(row=r, column=4))
ws_offers.column_dimensions["A"].width = 24
ws_offers.column_dimensions["B"].width = 15
ws_offers.column_dimensions["C"].width = 14
ws_offers.column_dimensions["D"].width = 14
ws_offers.column_dimensions["E"].width = 42
ws_offers.column_dimensions["F"].width = 50


# GTM sheet
set_title(ws_gtm, "Go-to-Market et Funnel", "Un entonnoir de conversion pragmatique : attention -> audit -> sprint -> projet -> retainer.")
for c, h in enumerate(["Canal", "Objectif", "Activite mensuelle", "Taux Prospect", "Taux Appel", "Taux Closing (Audit)", "Taux Signature", "Commentaire"], start=1):
    ws_gtm.cell(row=4, column=c, value=h)
style_header(ws_gtm, "A4:H4")
for r, row in enumerate(gtm_rows, start=5):
    channel, purpose, activity, ctr, lead, book, close, note = row
    ws_gtm.cell(row=r, column=1, value=channel)
    ws_gtm.cell(row=r, column=2, value=purpose)
    ws_gtm.cell(row=r, column=3, value=activity)
    ws_gtm.cell(row=r, column=4, value=ctr / 100 if ctr > 1 else ctr)
    ws_gtm.cell(row=r, column=5, value=lead / 100)
    ws_gtm.cell(row=r, column=6, value=book / 100)
    ws_gtm.cell(row=r, column=7, value=close / 100)
    ws_gtm.cell(row=r, column=8, value=note)
    fmt_pct(ws_gtm.cell(row=r, column=4))
    fmt_pct(ws_gtm.cell(row=r, column=5))
    fmt_pct(ws_gtm.cell(row=r, column=6))
    fmt_pct(ws_gtm.cell(row=r, column=7))
    if isinstance(activity, (int, float)):
        if activity > 1000:
            fmt_int(ws_gtm.cell(row=r, column=3))
        else:
            fmt_money(ws_gtm.cell(row=r, column=3))
ws_gtm["A11"] = "Constat pragmatique"
ws_gtm["B11"] = "Le modele de developpement ne depend pas d'un seul canal. LinkedIn etablit la confiance, les medias locaux la legitimite, les recommandations accelerent le cycle de vente et le demarchage direct comble les ecarts."
ws_gtm["B11"].alignment = Alignment(wrap_text=True)
ws_gtm.column_dimensions["A"].width = 26
ws_gtm.column_dimensions["B"].width = 36
ws_gtm.column_dimensions["C"].width = 16
ws_gtm.column_dimensions["D"].width = 12
ws_gtm.column_dimensions["E"].width = 12
ws_gtm.column_dimensions["F"].width = 14
ws_gtm.column_dimensions["G"].width = 12
ws_gtm.column_dimensions["H"].width = 34


# Revenue Model sheet
set_title(ws_rev, "Modele de Revenus", "Calcul mensuel attendu. Permet de visualiser l'evolution dans le temps de la capacite de facturation.")
ws_rev["A4"] = "Indicateur"
style_header(ws_rev, "A4:A4")
for m in range(HORIZON):
    c = 2 + m
    ws_rev.cell(row=4, column=c, value=month_label(m))
    ws_rev.cell(row=5, column=c, value=m + 1)
style_header(ws_rev, f"B4:{get_column_letter(1+HORIZON)}4")
style_header(ws_rev, f"B5:{get_column_letter(1+HORIZON)}5")

unit_start = 7
rev_start = 15
cogs_start = 23
row_map = {}
for i, o in enumerate(offers):
    unit_row = unit_start + i
    rev_row = rev_start + i
    cogs_row = cogs_start + i
    row_map[o["name"]] = (unit_row, rev_row, cogs_row)
    ws_rev.cell(row=unit_row, column=1, value=f"{o['name']} (Unites)")
    ws_rev.cell(row=rev_row, column=1, value=f"{o['name']} (CA)")
    ws_rev.cell(row=cogs_row, column=1, value=f"{o['name']} (Cout direct)")

    dr = driver_row_map[o["name"]]
    orow = offer_row_map[o["name"]]
    for m in range(HORIZON):
      col = 2 + m
      ucell = ws_rev.cell(row=unit_row, column=col)
      rcell = ws_rev.cell(row=rev_row, column=col)
      ccell = ws_rev.cell(row=cogs_row, column=col)
      col_letter = get_column_letter(col)
      ucell.value = f"=MIN(Assumptions!$D${dr},Assumptions!$B${dr}+(({col}-2)*Assumptions!$C${dr}))"
      rcell.value = f"={col_letter}{unit_row}*Assumptions!$B${orow}"
      ccell.value = f"={col_letter}{rev_row}*Assumptions!$C${orow}"
      fmt_money(rcell); fmt_money(ccell)
      fmt_pct(ws_rev.cell(row=unit_row, column=col))

tot_row = 32
labels = ["Chiffre d'Affaires global", "Cout direct total", "Marge brute (Valeur)", "Marge brute (%)"]
for idx, label in enumerate(labels):
    ws_rev.cell(row=tot_row + idx, column=1, value=label)
for m in range(HORIZON):
    col = 2 + m
    col_letter = get_column_letter(col)
    ws_rev.cell(row=tot_row, column=col, value=f"=SUM({col_letter}{rev_start}:{col_letter}{rev_start+len(offers)-1})")
    ws_rev.cell(row=tot_row + 1, column=col, value=f"=SUM({col_letter}{cogs_start}:{col_letter}{cogs_start+len(offers)-1})")
    ws_rev.cell(row=tot_row + 2, column=col, value=f"={col_letter}{tot_row}-{col_letter}{tot_row+1}")
    ws_rev.cell(row=tot_row + 3, column=col, value=f"=IFERROR({col_letter}{tot_row+2}/{col_letter}{tot_row},0)")
    fmt_money(ws_rev.cell(row=tot_row, column=col))
    fmt_money(ws_rev.cell(row=tot_row + 1, column=col))
    fmt_money(ws_rev.cell(row=tot_row + 2, column=col))
    fmt_pct(ws_rev.cell(row=tot_row + 3, column=col))

for r in range(1, tot_row + 4):
    ws_rev.cell(row=r, column=1).font = Font(bold=True) if r in (tot_row, tot_row + 1, tot_row + 2, tot_row + 3) else Font()
ws_rev.freeze_panes = "B6"
ws_rev.column_dimensions["A"].width = 28


# Expenses
set_title(ws_exp, "Charges de fonctionnement", "Details de l'Opex fixe par phase temporelle pour maintenir la discipline financiere.")
for c, h in enumerate(["Categorie d'Opex", "Mois 1-12", "Mois 13-24", "Mois 25-36"], start=1):
    ws_exp.cell(row=4, column=c, value=h)
style_header(ws_exp, "A4:D4")
exp_row_map = {}
for r, (label, p1, p2, p3) in enumerate(opex_rows, start=5):
    exp_row_map[label] = r
    ws_exp.cell(row=r, column=1, value=label)
    ws_exp.cell(row=r, column=2, value=p1)
    ws_exp.cell(row=r, column=3, value=p2)
    ws_exp.cell(row=r, column=4, value=p3)
    fmt_money(ws_exp.cell(row=r, column=2)); fmt_money(ws_exp.cell(row=r, column=3)); fmt_money(ws_exp.cell(row=r, column=4))
tot_opex_row = 5 + len(opex_rows) + 1
ws_exp.cell(row=tot_opex_row, column=1, value="Total charges fixes (Opex)")
for m in range(HORIZON):
    col = 2 + m
    col_letter = get_column_letter(col)
    # override with explicit formula using assumptions so changes on assumptions sheet cascade
    ws_exp.cell(row=tot_opex_row, column=col, value="=" + "+".join([f"IF(COLUMN()-1<=12,Assumptions!$B${opex_row_map[label]},IF(COLUMN()-1<=24,Assumptions!$C${opex_row_map[label]},Assumptions!$D${opex_row_map[label]}))" for label, *_ in opex_rows]))
    fmt_money(ws_exp.cell(row=tot_opex_row, column=col))
ws_exp.freeze_panes = "B5"
ws_exp.column_dimensions["A"].width = 42


# P&L
set_title(ws_pnl, "Compte de Resultat", "Vue previsionnelle mensuelle et synthese annuelle. Indispensable pour la credibilite du dossier.")
for m in range(HORIZON):
    ws_pnl.cell(row=4, column=2+m, value=month_label(m))
style_header(ws_pnl, f"B4:{get_column_letter(1+HORIZON)}4")
rows_pnl = {
    "Chiffre d'Affaires": 6,
    "Cout direct": 7,
    "Marge brute": 8,
    "Marge brute %": 9,
    "Frais de fonctionnement (Opex)": 10,
    "EBITDA": 11,
    "Impots (IS)": 12,
    "Resultat net": 13,
    "Marge EBITDA %": 14,
}
for label, row in rows_pnl.items():
    ws_pnl.cell(row=row, column=1, value=label)
    ws_pnl.cell(row=row, column=1).font = Font(bold=True)
for m in range(HORIZON):
    col = 2 + m
    col_letter = get_column_letter(col)
    ws_pnl.cell(row=6, column=col, value=f"='Revenue_Model'!{col_letter}{tot_row}")
    ws_pnl.cell(row=7, column=col, value=f"='Revenue_Model'!{col_letter}{tot_row+1}")
    ws_pnl.cell(row=8, column=col, value=f"='Revenue_Model'!{col_letter}{tot_row+2}")
    ws_pnl.cell(row=9, column=col, value=f"=IFERROR({col_letter}8/{col_letter}6,0)")
    ws_pnl.cell(row=10, column=col, value=f"='Expenses'!{col_letter}{tot_opex_row}")
    ws_pnl.cell(row=11, column=col, value=f"={col_letter}8-{col_letter}10")
    ws_pnl.cell(row=12, column=col, value=f"=MAX({col_letter}11,0)*Assumptions!$B$8")
    ws_pnl.cell(row=13, column=col, value=f"={col_letter}11-{col_letter}12")
    ws_pnl.cell(row=14, column=col, value=f"=IFERROR({col_letter}11/{col_letter}6,0)")
    for row in [6,7,8,10,11,12,13]:
        fmt_money(ws_pnl.cell(row=row, column=col))
    for row in [9,14]:
        fmt_pct(ws_pnl.cell(row=row, column=col))

annual_row = 17
ws_pnl.cell(row=annual_row, column=1, value="Synthese annuelle")
ws_pnl.cell(row=annual_row, column=1).font = Font(bold=True)
for j, year_label in enumerate(["Annee 1", "Annee 2", "Annee 3"], start=2):
    ws_pnl.cell(row=annual_row, column=j, value=year_label)
style_header(ws_pnl, f"B{annual_row}:D{annual_row}")
year_ranges = [(2, 14), (14, 26), (26, 38)]
for idx, (c1, c2) in enumerate(year_ranges, start=2):
    letter1 = get_column_letter(c1)
    letter2 = get_column_letter(c2-1)
    ws_pnl.cell(row=annual_row + 1, column=1, value="Chiffre d'Affaires")
    ws_pnl.cell(row=annual_row + 2, column=1, value="Marge brute")
    ws_pnl.cell(row=annual_row + 3, column=1, value="EBITDA")
    ws_pnl.cell(row=annual_row + 4, column=1, value="Resultat net")
    ws_pnl.cell(row=annual_row + 5, column=1, value="Tresorerie finale")
    ws_pnl.cell(row=annual_row + 1, column=idx, value=f"=SUM('{ws_pnl.title}'!{letter1}6:{letter2}6)")
    ws_pnl.cell(row=annual_row + 2, column=idx, value=f"=SUM('{ws_pnl.title}'!{letter1}8:{letter2}8)")
    ws_pnl.cell(row=annual_row + 3, column=idx, value=f"=SUM('{ws_pnl.title}'!{letter1}11:{letter2}11)")
    ws_pnl.cell(row=annual_row + 4, column=idx, value=f"=SUM('{ws_pnl.title}'!{letter1}13:{letter2}13)")
    ws_pnl.cell(row=annual_row + 5, column=idx, value=f"='Cashflow'!{get_column_letter(c2-1)}12")
    for row in range(annual_row + 1, annual_row + 6):
        fmt_money(ws_pnl.cell(row=row, column=idx))
ws_pnl.freeze_panes = "B5"
ws_pnl.column_dimensions["A"].width = 28


# Cashflow
set_title(ws_cash, "Flux de Tresorerie", "Flux de tresorerie reels. C'est l'indicateur de securite et de survie que l'investisseur scrute apres les profits.")
for m in range(HORIZON):
    ws_cash.cell(row=4, column=2+m, value=month_label(m))
style_header(ws_cash, f"B4:{get_column_letter(1+HORIZON)}4")
cash_rows = {
    "Solde initial": 6,
    "EBITDA": 7,
    "Impots payes": 8,
    "Investissements (Capex)": 9,
    "Apport financement": 10,
    "Flux net de tresorerie": 11,
    "Solde final": 12,
}
for label, row in cash_rows.items():
    ws_cash.cell(row=row, column=1, value=label)
    ws_cash.cell(row=row, column=1).font = Font(bold=True)
for m in range(HORIZON):
    col = 2 + m
    col_letter = get_column_letter(col)
    if m == 0:
        ws_cash.cell(row=6, column=col, value="=Assumptions!$B$9")
    else:
        prev_col = get_column_letter(col - 1)
        ws_cash.cell(row=6, column=col, value=f"={prev_col}12")
    ws_cash.cell(row=7, column=col, value=f"='P&L'!{col_letter}11")
    ws_cash.cell(row=8, column=col, value=f"='P&L'!{col_letter}12")
    # FIX: replaced single quotes around dates with double quotes to resolve parse errors in Google Sheets/Excel
    ws_cash.cell(row=9, column=col, value=f'=IF({col_letter}$4="{month_label(0)}",Assumptions!$B$11,IF({col_letter}$4="{month_label(12)}",20000,IF({col_letter}$4="{month_label(24)}",15000,0)))')
    ws_cash.cell(row=10, column=col, value="0")
    ws_cash.cell(row=11, column=col, value=f"={col_letter}7-{col_letter}8-{col_letter}9+{col_letter}10")
    ws_cash.cell(row=12, column=col, value=f"={col_letter}6+{col_letter}11")
    for row in range(6, 13):
        fmt_money(ws_cash.cell(row=row, column=col))
ws_cash.freeze_panes = "B5"
ws_cash.column_dimensions["A"].width = 24


# Funding
set_title(ws_fund, "Plan de Financement", "Deux visions : developpement commando (autofinance) et developpement accelere (levée de fonds).")
fund_headers = ["Poste", "Montant / Part", "Remarques"]
for c, h in enumerate(fund_headers, start=1):
    ws_fund.cell(row=4, column=c, value=h)
style_header(ws_fund, "A4:C4")
fund_items = [
    ("Besoin minimum (Runway lean)", 300_000, "Suffisant si pilote par le fondateur et livraison stricte."),
    ("Financement recommande (Seed)", SEED_TARGET, "Capital d'acceleration : force commerciale, experts, IP et reserve."),
    ("Affectation : Ventes & Image de marque", 0.25, "Creation de contenu, visibilite, acquisition et conversion."),
    ("Affectation : Capacite de livraison", 0.30, "Sous-traitants, support technique et vitesse d'execution."),
    ("Affectation : Standardisation & IP", 0.20, "Modeles d'agents, outils internes, capitalisation et architecture."),
    ("Affectation : Conformite & Administration", 0.10, "Fiscalite locale, contrats, structure et standards juridiques."),
    ("Affectation : Reserve de roulement", 0.15, "Securiser l'entreprise face aux delais de paiement."),
]
for r, (item, value, note) in enumerate(fund_items, start=5):
    ws_fund.cell(row=r, column=1, value=item)
    ws_fund.cell(row=r, column=2, value=value)
    ws_fund.cell(row=r, column=3, value=note)
    ws_fund.cell(row=r, column=1).font = Font(bold=True)
    if isinstance(value, float) and value < 1:
        fmt_pct(ws_fund.cell(row=r, column=2))
    else:
        fmt_money(ws_fund.cell(row=r, column=2))
ws_fund["A14"] = "Note stratégique"
ws_fund["B14"] = "L'entreprise n'a pas besoin de fonds demesures pour exister. La levee de fonds sert uniquement a accelerer la commercialisation et a proteger le temps du fondateur pour fiabiliser la livraison."
ws_fund["B14"].alignment = Alignment(wrap_text=True)
ws_fund.column_dimensions["A"].width = 38
ws_fund.column_dimensions["B"].width = 18
ws_fund.column_dimensions["C"].width = 60


# Scenarios
set_title(ws_scn, "Analyse des Scenarios", "Scénarios conservateurs, de base et agressifs pour tester la resistance du modele.")
for c, h in enumerate(["Indicateur", "Conservateur", "Reference (Base)", "Optimiste"], start=1):
    ws_scn.cell(row=4, column=c, value=h)
style_header(ws_scn, "A4:D4")
scenario_metrics = [
    ("CA - Annee 1", total(conservative["rev"][:12]), total(base["rev"][:12]), total(aggressive["rev"][:12])),
    ("CA - Annee 2", total(conservative["rev"][12:24]), total(base["rev"][12:24]), total(aggressive["rev"][12:24])),
    ("CA - Annee 3", total(conservative["rev"][24:]), total(base["rev"][24:]), total(aggressive["rev"][24:])),
    ("EBITDA - Annee 1", total(conservative["ebitda"][:12]), total(base["ebitda"][:12]), total(aggressive["ebitda"][:12])),
    ("EBITDA - Annee 2", total(conservative["ebitda"][12:24]), total(base["ebitda"][12:24]), total(aggressive["ebitda"][12:24])),
    ("EBITDA - Annee 3", total(conservative["ebitda"][24:]), total(base["ebitda"][24:]), total(aggressive["ebitda"][24:])),
    ("Tresorerie finale (Mois 36)", conservative["cash"][-1], base["cash"][-1], aggressive["cash"][-1]),
]
for r, rowv in enumerate(scenario_metrics, start=5):
    ws_scn.cell(row=r, column=1, value=rowv[0])
    ws_scn.cell(row=r, column=2, value=rowv[1])
    ws_scn.cell(row=r, column=3, value=rowv[2])
    ws_scn.cell(row=r, column=4, value=rowv[3])
    for c in range(2, 5):
        fmt_money(ws_scn.cell(row=r, column=c))
ws_scn["A14"] = "Constats sur le modele de base :"
ws_scn["A15"] = f"Mois d'equilibre (seuil) : {next((i+1) for i,v in enumerate(base['ebitda']) if v > 0)}"
ws_scn["A16"] = f"Marge brute moyenne (annee 1) : {total(base['gross'][:12]) / max(total(base['rev'][:12]),1):.1%}"
ws_scn["A17"] = f"Solde de tresorerie final (mois 36) : ZAR {base['cash'][-1]:,.0f}"
ws_scn["A18"] = "Le scenario conservateur majore l'Opex de 5% et minore les revenus attendus de 18%."
ws_scn["A19"] = "Le scenario optimiste augmente les ventes de 18% avec un Opex ajuste a +10% pour absorber le scale."
ws_scn.column_dimensions["A"].width = 38
ws_scn.column_dimensions["B"].width = 16
ws_scn.column_dimensions["C"].width = 16
ws_scn.column_dimensions["D"].width = 16


# Risks / Q&A
set_title(ws_risk, "Gestion des Risques & Q&A", "Questions et objections susceptibles d'etre posees par des partenaires ou investisseurs.")
for c, h in enumerate(["Type", "Sujet", "Risque / Question", "Response / Mitigation", "Responsable"], start=1):
    ws_risk.cell(row=4, column=c, value=h)
style_header(ws_risk, "A4:E4")
row = 5
for item in risks:
    ws_risk.cell(row=row, column=1, value="Risque")
    ws_risk.cell(row=row, column=2, value=item[0])
    ws_risk.cell(row=row, column=3, value=item[0])
    ws_risk.cell(row=row, column=4, value=f"Probabilite : {item[1]} | Declencheur : {item[2]} | Plan d'action : {item[3]}")
    ws_risk.cell(row=row, column=5, value=item[4])
    row += 1
for item in qa:
    ws_risk.cell(row=row, column=1, value="Q&A")
    ws_risk.cell(row=row, column=2, value=item[0])
    ws_risk.cell(row=row, column=3, value=item[0])
    ws_risk.cell(row=row, column=4, value=item[1])
    ws_risk.cell(row=row, column=5, value="Opays")
    row += 1
for r in range(5, row):
    for c in range(1, 6):
        ws_risk.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
ws_risk.column_dimensions["A"].width = 12
ws_risk.column_dimensions["B"].width = 34
ws_risk.column_dimensions["C"].width = 34
ws_risk.column_dimensions["D"].width = 72
ws_risk.column_dimensions["E"].width = 14


# Sources
set_title(ws_src, "Sources de Conformite", "Sources officielles utilisees pour formuler les hypotheses de taxes et de fiscalite.")
for c, h in enumerate(["Sujet", "Source officielle", "Lien (URL)"], start=1):
    ws_src.cell(row=4, column=c, value=h)
style_header(ws_src, "A4:C4")
src_rows = [
    ("TVA", "Page officielle de la SARS sur la TVA", "https://www.sars.gov.za/types-of-tax/value-added-tax/"),
    ("Impot sur les societes", "Taux standard d'IS de la SARS", "https://www.sars.gov.za/types-of-tax/corporate-income-tax/"),
    ("Exclusion taxe chiffre d'affaires", "Conditions de la SARS pour les services professionnels", "https://www.sars.gov.za/faq/faq-what-are-the-requirements-for-a-micro-business-to-qualify-for-turnover-tax/"),
]
for r, rowv in enumerate(src_rows, start=5):
    for c, val in enumerate(rowv, start=1):
        ws_src.cell(row=r, column=c, value=val)
    for c in range(1, 4):
        ws_src.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
ws_src.column_dimensions["A"].width = 34
ws_src.column_dimensions["B"].width = 34
ws_src.column_dimensions["C"].width = 72


# Operations
set_title(ws_ops, "Systeme d'Exploitation", "C'est l'execution et le delivery standardise qui creent la valeur du cabinet, pas seulement les idees.")
ops_sections = [
    ("Principes d'operation", [
        "Une seule source de verite pour la strategie, la livraison et les decisions.",
        "Des ajustements petits et tracables plutot que de grandes improvisations.",
        "Chaque mission client a un debut, une livraison et une sortie definis.",
        "Le fondateur se concentre sur les ventes et la qualite, pas sur les taches repetitives.",
    ]),
    ("Core roles", [
        "CEO / Fondateur : ventes, narratif, design strategique, validation finale de qualite.",
        "Responsable de la Livraison : coordination des projets, respect des delais, relation client.",
        "Recherche & R&D : travaux scientifiques, documents, analyses et synthesis.",
        "Operations & Admin : planification, facturation, suivi administratif et hygiene documentaire.",
    ]),
    ("Regles de decision", [
        "Si cela ne cree pas de valeur ou ne reduit pas la friction, ce n'est pas prioritaire.",
        "Si le perimetre (scope) change, faire un avenant ou reporter.",
        "Si la livraison menace la marge, refondre l'offre commerciale.",
        "Si un processus se repete 3 fois, il doit etre standardise.",
    ]),
    ("Cadence d'execution", [
        "Quotidien : revue des opportunites, bloc de livraison, bloc de suivi.",
        "Hebdomadaire : revue du pipeline, revue de livraison, revue de tresorerie, retour d'experience.",
        "Mensuel : revue des indicateurs (KPIs), revue des prix, revue de capacite, mise a jour du journal de decisions.",
    ]),
]
row = 4
for title, items in ops_sections:
    ws_ops.cell(row=row, column=1, value=title)
    ws_ops.cell(row=row, column=1).font = Font(bold=True)
    ws_ops.cell(row=row, column=1).fill = blue_fill
    row += 1
    for item in items:
        ws_ops.cell(row=row, column=1, value=f"• {item}")
        ws_ops.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1
ws_ops["D4"] = "Facteur clé de succes"
ws_ops["D5"] = "L'entreprise est rentable quand le message, le processus, le prix et le rythme de livraison pointent tous dans la meme direction."
ws_ops["D6"] = "Les investisseurs s'assurent que le cabinet peut livrer de la valeur sans reposer uniquement sur un effort heroique du fondateur."
ws_ops["D7"] = "Le but n'est pas la complexite technique, mais l'assurance d'une execution previsible."
for r in range(4, 8):
    ws_ops[f"D{r}"].alignment = Alignment(wrap_text=True)
ws_ops.column_dimensions["A"].width = 72
ws_ops.column_dimensions["D"].width = 48


# Capacity
set_title(ws_cap, "Capacite de l'Equipe", "Logique de debits. Ce modele garde l'entreprise honnete sur ce qu'elle est en mesure de livrer physiquement.")
for c, h in enumerate(["Service", "Heures Fondateur", "Heures Expert", "Heures Ops", "Total Heures", "Heures disponibles / mois", "Unites max / mois"], start=1):
    ws_cap.cell(row=4, column=c, value=h)
style_header(ws_cap, "A4:G4")
capacity_rows = [
    ("Audit / Diagnostic", 12, 4, 4, 20, 80),
    ("Sprint d'Efficience", 24, 12, 8, 44, 100),
    ("Projet d'Implementation", 60, 30, 15, 105, 160),
    ("Forge Retainer (Accompagnement)", 6, 3, 3, 12, 90),
    ("Dossier Sovereign", 30, 20, 10, 60, 70),
]
for r, rowv in enumerate(capacity_rows, start=5):
    service, founder_h, spec_h, ops_h, total_h, available_h = rowv
    ws_cap.cell(row=r, column=1, value=service)
    ws_cap.cell(row=r, column=2, value=founder_h)
    ws_cap.cell(row=r, column=3, value=spec_h)
    ws_cap.cell(row=r, column=4, value=ops_h)
    ws_cap.cell(row=r, column=5, value=total_h)
    ws_cap.cell(row=r, column=6, value=available_h)
    ws_cap.cell(row=r, column=7, value=f"=F{r}/E{r}")
    fmt_int(ws_cap.cell(row=r, column=2))
    fmt_int(ws_cap.cell(row=r, column=3))
    fmt_int(ws_cap.cell(row=r, column=4))
    fmt_int(ws_cap.cell(row=r, column=5))
    fmt_int(ws_cap.cell(row=r, column=6))
    fmt_pct(ws_cap.cell(row=r, column=7))
ws_cap["A12"] = "Interpretation"
ws_cap["B12"] = "Les plafonds de capacite sont volontairement conservateurs ; ils empechent le modele de revenus de presumer d'une bande passante de livraison infinie."
ws_cap["B12"].alignment = Alignment(wrap_text=True)
ws_cap.column_dimensions["A"].width = 26
ws_cap.column_dimensions["B"].width = 16
ws_cap.column_dimensions["C"].width = 18
ws_cap.column_dimensions["D"].width = 14
ws_cap.column_dimensions["E"].width = 12
ws_cap.column_dimensions["F"].width = 22
ws_cap.column_dimensions["G"].width = 16


# Delivery
set_title(ws_del, "Parcours client (Delivery)", "La qualite de livraison de nos services fait notre reputation. Ce playbook structure chaque phase.")
for c, h in enumerate(["Etape", "Responsable", "Criteres d'entree", "Criteres de sortie", "Livrable", "Duree standard"], start=1):
    ws_del.cell(row=4, column=c, value=h)
style_header(ws_del, "A4:F4")
delivery_flow = [
    ("Opportunite recue", "Fondateur", "Entrant, recommandation ou reponse a la prospection", "Qualifie ou disqualifie", "Opportunite enregistree dans le CRM", "1 jour"),
    ("Appel de decouverte", "Fondateur", "Besoin reel et budget existant", "Probleme et objectifs compris", "Notes d'appel + prochaine action", "30-45 min"),
    ("Audit / Diagnostic", "Fondateur + R&D", "Le client accepte l'offre d'entree", "Cartographie de friction et hypothese de ROI pretes", "Note d'audit et de synthese", "3-10 jours"),
    ("Proposition / Contrat", "Fondateur", "Diagnostic complete", "Perimetre, tarif et planning valides", "Proposition commerciale + contrat signe", "1-3 jours"),
    ("Sprint de livraison", "Responsable Livraison", "Contrat signe et acompte verse", "Flux de travail ou systeme cible ameliore", "Dossier d'implementation standard", "2-6 semaines"),
    ("Transfert final", "Fondateur + Ops", "Criteres d'acceptation valides", "Le client sait utiliser le nouveau processus", "Note de transfert + procedure standard", "1-2 jours"),
    ("Abonnement (Retainer)", "Fondateur", "Valeur pouvee et confirmee", "Accord d'optimisation recurrente signe", "Point mensuel de suivi operationnel", "Continu"),
]
for r, rowv in enumerate(delivery_flow, start=5):
    for c, val in enumerate(rowv, start=1):
        ws_del.cell(row=r, column=c, value=val)
    for c in range(1, 7):
        ws_del.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
ws_del["A14"] = "Regle d'or"
ws_del["B14"] = "Definir les criteres de reussite en amont du projet. Ne jamais declarer un projet termine avant que le client n'ait valide la prise en main de l'outil."
ws_del["B14"].alignment = Alignment(wrap_text=True)
ws_del.column_dimensions["A"].width = 24
ws_del.column_dimensions["B"].width = 18
ws_del.column_dimensions["C"].width = 28
ws_del.column_dimensions["D"].width = 28
ws_del.column_dimensions["E"].width = 30
ws_del.column_dimensions["F"].width = 18


# Cadence
set_title(ws_cad, "Rhythme Hebdomadaire", "Un rythme d'execution standardise pour transformer la vision en routine.")
for c, h in enumerate(["Jour", "Focus", "Actions cles", "Livrable"], start=1):
    ws_cad.cell(row=4, column=c, value=h)
style_header(ws_cad, "A4:D4")
cadence_rows = [
    ("Lundi", "Pipeline et priorites", "Revue des opportunites, planification client, definition des 3 objectifs de la semaine", "Plan d'action clair et responsabilites attribuees"),
    ("Mardi", "Livraison / Production", "Concentration sur les livrables clients et les modeles internes", "Produits livrables et shippable"),
    ("Mercredi", "Ventes & Visibilite", "Demarchage, LinkedIn, presse locale, relations publiques", "Activites de prospection"),
    ("Jeudi", "Livraison & Echanges client", "Points d'avancement avec les clients, suivi, QA et validations", "Confiance et dynamique client"),
    ("Vendredi", "Finance, Apprentissage & Admin", "Revue de tresorerie, facturation, journal des decisions, retour d'experience", "Cloture propre et preparation de la semaine suivante"),
]
for r, rowv in enumerate(cadence_rows, start=5):
    for c, val in enumerate(rowv, start=1):
        ws_cad.cell(row=r, column=c, value=val)
    for c in range(1, 5):
        ws_cad.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
ws_cad["A12"] = "Regle d'or"
ws_cad["B12"] = "Si la semaine ne produit pas d'avancement sur le pipeline, sur la livraison ou sur la tresorerie, le rythme n'est pas bon."
ws_cad["B12"].alignment = Alignment(wrap_text=True)
ws_cad.column_dimensions["A"].width = 14
ws_cad.column_dimensions["B"].width = 22
ws_cad.column_dimensions["C"].width = 44
ws_cad.column_dimensions["D"].width = 26


# SOPs
set_title(ws_sop, "Procedures standards (SOPs)", "Les SOPs garantissent la repetabilite et la valeur a terme du cabinet.")
for c, h in enumerate(["Procedure (SOP)", "Objectif", "Declencheur", "Responsable", "Resultat attendu"], start=1):
    ws_sop.cell(row=4, column=c, value=h)
style_header(ws_sop, "A4:E4")
sops = [
    ("Qualification des leads", "Filtrer les opportunites viables", "Nouveau contact ou recommandation", "Fondateur", "Pipeline qualifie ou disqualification propre"),
    ("Appel de decouverte", "Comprendre le probleme reel", "Apres la qualification", "Fondateur", "Probleme defini et alignement de fit"),
    ("Checklist d'audit", "Standardiser nos diagnostics", "Le client achete un diagnostic", "Fondateur + R&D", "Note d'audit standardisee"),
    ("Modele de proposition", "Accelerer le suivi commercial", "Audit complete", "Fondateur", "Offre commerciale claire"),
    ("Controle du perimetre", "Proteger notre marge brute", "Toute demande de changement de perimetre", "Fondateur", "Avenant signe ou gel du perimetre"),
    ("Revue qualite (QA)", "Eviter les erreurs de livraison", "Avant le transfert final", "Responsable Livraison", "Livrable verifie et certifie"),
    ("Revue de tresorerie", "Proteger notre runway", "Chaque vendredi", "Fondateur + Ops", "Situation de tresorerie a jour et actions"),
    ("Journal de decisions", "Maintenir une gouvernance vivante", "Toute decision strategique", "Fondateur", "Decision consignee et datee"),
]
for r, rowv in enumerate(sops, start=5):
    for c, val in enumerate(rowv, start=1):
        ws_sop.cell(row=r, column=c, value=val)
    for c in range(1, 6):
        ws_sop.cell(row=r, column=c).alignment = Alignment(wrap_text=True)
ws_sop["A15"] = "Regle d'or"
ws_sop["B15"] = "Si une action complexe est executee plus d'une fois, elle doit etre documentee par une SOP."
ws_sop["B15"].alignment = Alignment(wrap_text=True)
ws_sop.column_dimensions["A"].width = 24
ws_sop.column_dimensions["B"].width = 26
ws_sop.column_dimensions["C"].width = 22
ws_sop.column_dimensions["D"].width = 16
ws_sop.column_dimensions["E"].width = 34


# Team
set_title(ws_team, "Modele d'Equipe", "L'entreprise doit etre visible et quantifiable comme un collectif et non seulement par son fondateur.")
for c, h in enumerate(["Role", "Responsabilite principale", "Statut", "Capacite (heures/mois)", "Repartition hebdomadaire", "Declencheur de recrutement"], start=1):
    ws_team.cell(row=4, column=c, value=h)
style_header(ws_team, "A4:F4")
team_rows = [
    ("Fénelon Lamsasiri — CEO / R&D", "Vision, gouvernance, strategie, validation qualite, recherche", "Actif", 80, "Strategie 30%, livraison 25%, management 25%, contenu 20%", "Lorsque la strategie ou la qualite depasse 80h/mois"),
    ("Prince Bagheni — Directeur Commercial", "Prospection, closing, relations partenaires, rigueur du pipeline", "Actif", 70, "Prospection 45%, suivi 30%, closing 15%, reporting 10%", "Lorsque le volume d'opportunites requiert un commercial dedie"),
    ("Patricia Zamwana — Admin & Gestion", "Support prospection, suivi comptable, tresorerie, soutien administratif", "Actif", 60, "Finance 40%, admin 30%, support ventes 30%", "Lorsque la facturation ou le suivi de tresorerie requiert un temps plein"),
    ("Zaina Bwale Godlove — Marque & Ventes", "Prospection, communication, visibilite, supports institutionnels", "Actif", 60, "Comms 40%, prospection 30%, marque 20%, support 10%", "Lorsque l'effort de communication ou de visibilite augmente"),
    ("Directeur Technique / CTO", "Architecture, securite, qualite technique, supervision livraison", "Vacant / En recrutement", 0, "N/A", "Lorsque la livraison technique requiert un responsable dedie"),
    ("Antigravity — IA Collaboratrice", "Redaction, structuration, support automatisation, aide operationnelle", "Actif (actif de type systeme)", 40, "Documents 30%, automatisation 30%, support 40%", "Lorsque les taches administratives ou repetitives saturent l'equipe"),
]
for r, rowv in enumerate(team_rows, start=5):
    for c, val in enumerate(rowv, start=1):
        ws_team.cell(row=r, column=c, value=val)
    for c in range(1, 7):
        ws_team.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
ws_team["A12"] = "Regle d'equipe"
ws_team["B12"] = "L'entreprise grandit en capitalisant sur ses process avant d'ajouter de la masse salariale. Chaque recrutement doit soit securiser la marge, soit augmenter le debit de livraison, soit securiser le temps du fondateur."
ws_team["B12"].alignment = Alignment(wrap_text=True)
ws_team.column_dimensions["A"].width = 20
ws_team.column_dimensions["B"].width = 34
ws_team.column_dimensions["C"].width = 18
ws_team.column_dimensions["D"].width = 16
ws_team.column_dimensions["E"].width = 34
ws_team.column_dimensions["F"].width = 34


# Pipeline
set_title(ws_pipe, "Modélisation du Pipeline", "Liaison mathematique entre l'activite de prospection, les leads et le chiffre d'affaires.")
for c, h in enumerate(["Canal", "Activite mensuelle", "Taux Prospect", "Taux Decouverte", "Taux Audit", "Taux Closing", "Offre par defaut", "Valeur mensuelle attendue"], start=1):
    ws_pipe.cell(row=4, column=c, value=h)
style_header(ws_pipe, "A4:H4")
pipeline_rows = [
    ("Page entreprise LinkedIn", 8, 0.12, 0.35, 0.55, 0.30, "Audit / Diagnostic"),
    ("Presse locale / communiques", 2, 0.30, 0.50, 0.60, 0.35, "Sprint d'Efficience"),
    ("Recommandations / Partenaires", 4, 0.45, 0.60, 0.70, 0.50, "Projet d'Implementation"),
    ("Prospection directe", 40, 0.04, 0.25, 0.40, 0.18, "Audit / Diagnostic"),
]
for r, rowv in enumerate(pipeline_rows, start=5):
    channel, activity, lead_rate, disc_rate, audit_rate, close_rate, default_offer = rowv
    ws_pipe.cell(row=r, column=1, value=channel)
    ws_pipe.cell(row=r, column=2, value=activity)
    ws_pipe.cell(row=r, column=3, value=lead_rate)
    ws_pipe.cell(row=r, column=4, value=disc_rate)
    ws_pipe.cell(row=r, column=5, value=audit_rate)
    ws_pipe.cell(row=r, column=6, value=close_rate)
    ws_pipe.cell(row=r, column=7, value=f"=B{r}*C{r}*D{r}*E{r}")
    ws_pipe.cell(row=r, column=8, value=f"=G{r}*F{r}*Assumptions!$B${offer_row_map[default_offer]}")
    fmt_pct(ws_pipe.cell(row=r, column=3))
    fmt_pct(ws_pipe.cell(row=r, column=4))
    fmt_pct(ws_pipe.cell(row=r, column=5))
    fmt_pct(ws_pipe.cell(row=r, column=6))
    fmt_money(ws_pipe.cell(row=r, column=8))
    for c in range(1, 9):
        ws_pipe.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
ws_pipe["A12"] = "Regle d'or"
ws_pipe["B12"] = "Le modele prevoir n'a de valeur que si les conversions refletent des discussions reelles. Ajustez les taux et l'activite mensuellement."
ws_pipe["B12"].alignment = Alignment(wrap_text=True)
ws_pipe["A16"] = "Canaux de communication actuels :"
ws_pipe["B16"] = "LinkedIn : https://www.linkedin.com/company/opaystech/ | X : https://x.com/OpaysTech | Facebook : https://www.facebook.com/share/1HXhTPzytG/?mibextid=wwXlfr | Site : https://opays.io/"
ws_pipe["B16"].alignment = Alignment(wrap_text=True)
ws_pipe["A14"] = "Valeur mensuelle estimee du pipeline"
ws_pipe["B14"] = "=SUM(H5:H8)"
fmt_money(ws_pipe["B14"])
ws_pipe.column_dimensions["A"].width = 26
ws_pipe.column_dimensions["B"].width = 16
ws_pipe.column_dimensions["C"].width = 12
ws_pipe.column_dimensions["D"].width = 14
ws_pipe.column_dimensions["E"].width = 12
ws_pipe.column_dimensions["F"].width = 12
ws_pipe.column_dimensions["G"].width = 16
ws_pipe.column_dimensions["H"].width = 22


# 90 day plan
set_title(ws_90, "Plan d'action 90 jours", "Feuille de route pour le premier trimestre afin d'aligner l'effort commercial et la livraison.")
for c, h in enumerate(["Semaine", "Objectif principal", "Actions cles", "Responsable", "Indicateur (KPI)", "Resultat attendu"], start=1):
    ws_90.cell(row=4, column=c, value=h)
style_header(ws_90, "A4:F4")
weeks = [
    ("S1", "Fixer la reference", "Revoir les offres, valider le pipeline, confirmer les roles, combler les ecarts documentaires", "Fenelon + Prince", "Modele et suivi approuves", "Base de reference operationnelle unique"),
    ("S2", "Affiner l'offre", "Ajuster les prix, eliminer l'ambiguite de perimetre, finaliser les descriptions d'offres", "Prince + Fenelon", "3 offres formalisees", "Catalogue commercial clair et fige"),
    ("S3", "Batir le pipeline", "Lancer la liste de prospection, cadence LinkedIn, communiques presse, contacts partenaires", "Prince", "20 prospects qualifies", "Liste de pipeline active"),
    ("S4", "Rigueur de livraison", "Mettre en place les SOPs, attribuer les roles de livraison, lancer la revue hebdo", "Responsable Livraison", "1ere livraison sur checklist", "Systeme de livraison etabli"),
    ("S5", "Preuves de valeur", "Rediger une etude de cas, un temoignage client, des mesures avant/apres", "Fenelon + R&D", "1 actif de preuve valide", "Outil de credibilite commerciale"),
    ("S6", "Conversion", "Transformer les audits en sprints et les sprints en propositions d'implementation", "Prince", "2 proposals sent", "Dynamique commerciale active"),
    ("S7", "Controle de capacite", "Auditer la charge par rapport a la capacite, decider du recours a des prestataires", "Fenelon + Ops", "Indicateurs de capacite mis a jour", "Gestion saine de la charge"),
    ("S8", "Dossier Investisseur v1", "Rediger la structure du pitch, mettre a jour les chiffres, valider les risques", "Fenelon", "Structure de pitch prete", "Narratif investisseur pret"),
    ("S9", "Preuve de repetition", "Repeter un deploiement reussi sur un second client ou prospect", "Responsable Livraison", "2eme etude de cas validee", "Preuve de repetabilite etablie"),
    ("S10", "Rigueur financiere", "Suivi des factures, depenses, runway, TVA et IS", "Ops + DG", "Suivi de tresorerie a jour", "Controle financier total"),
    ("S11", "Semaine de decision", "Confirmer ce qu'il faut standardiser, arreter, deleguer ou developper", "Fenelon", "Journal des decisions a jour", "Priorites strategiques nettes"),
    ("S12", "Revue trimestrielle", "Revue des KPIs, de la strategie, du mix d'offres et preparation du trimestre suivant", "Fenelon + Equipe", "Rapport trimestriel finalise", "Plan operationnel du trimestre suivant"),
]
for r, rowv in enumerate(weeks, start=5):
    for c, val in enumerate(rowv, start=1):
        ws_90.cell(row=r, column=c, value=val)
    for c in range(1, 7):
        ws_90.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
ws_90["A18"] = "Regle d'or"
ws_90["B18"] = "Le trimestre est reussi s'il accroit la qualite du pipeline, la tresorerie et la repetabilite, plutot que la simple activite."
ws_90["B18"].alignment = Alignment(wrap_text=True)
ws_90.column_dimensions["A"].width = 10
ws_90.column_dimensions["B"].width = 24
ws_90.column_dimensions["C"].width = 48
ws_90.column_dimensions["D"].width = 16
ws_90.column_dimensions["E"].width = 20
ws_90.column_dimensions["F"].width = 26


# Pitch outline
set_title(ws_pitch, "Structure du Pitch Deck", "Planification des 10 slides cles de presentation aux investisseurs et partenaires.")
for c, h in enumerate(["Slide", "Message cle", "Preuve / Contenu detaille"], start=1):
    ws_pitch.cell(row=4, column=c, value=h)
style_header(ws_pitch, "A4:C4")
pitch_rows = [
    ("1. Titre", "Opays Tech - L'ingenierie de l'efficience au service des operations et de l'adoption de l'IA.", "Marque, logo et positionnement en une phrase"),
    ("2. Probleme", "Les organisations perdent du temps, de la marge et de la confiance a cause de flux de travail disperses.", "Points de friction, exemples et contexte local"),
    ("3. Solution", "Forge execute et rationalise ; Sovereign batit l'autonomie et la resilience strategique.", "Presentation de notre modele hybride"),
    ("4. Pourquoi maintenant", "L'adoption de l'IA est desordonnee ; les clients exigent de l'execution pratique et mesurable.", "Opportunite de marche"),
    ("5. Offres & Marges", "Audit, sprint, projet d'implementation, abonnements (retainer) et dossiers.", "Echelle d'offres et logique de prix productises"),
    ("6. Traction & Canaux", "Lancement initie ; signaux d'interet reels sur LinkedIn et les canaux locaux.", "LinkedIn, presse locale et dynamique de pipeline"),
    ("7. Modele financier", "Des services a haute marge aujourd'hui pour financer la R&D de demain.", "Indicateurs financiers issus du previsionnel"),
    ("8. Modele operationnel", "Une execution rigoureuse appuyee par des SOPs, une cadence fixe et un controle de capacite.", "Integration de notre systeme interne"),
    ("9. Risques & Reponses", "Gestion de la dependance au fondateur, de la derive de perimetre et du rythme de conversion.", "Synthese de l'onglet Risques_QA"),
    ("10. Demande & Allocation", "Financement Seed pour le developpement commercial, la livraison, la conformite et le runway.", "Synthese de l'onglet Financement"),
]
for r, rowv in enumerate(pitch_rows, start=5):
    for c, val in enumerate(rowv, start=1):
        ws_pitch.cell(row=r, column=c, value=val)
    for c in range(1, 4):
        ws_pitch.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
ws_pitch["A17"] = "Regle d'or"
ws_pitch["B17"] = "Chaque diapositive doit repondre de maniere concise a une question essentielle d'un investisseur. Evitez de surcharger les visuels."
ws_pitch["B17"].alignment = Alignment(wrap_text=True)
ws_pitch.column_dimensions["A"].width = 20
ws_pitch.column_dimensions["B"].width = 54
ws_pitch.column_dimensions["C"].width = 54


# Investor Q&A
set_title(ws_qa, "Dossier Questions / Reponses", "Questions sensibles frequemment posees par les comites de selection ou investisseurs, avec reponses directes.")
for c, h in enumerate(["Question", "Pourquoi cela compte", "Reponse", "Preuve / Document associe"], start=1):
    ws_qa.cell(row=4, column=c, value=h)
style_header(ws_qa, "A4:D4")
qa_rows = [
    ("Pourquoi Opays maintenant ?", "Timing et urgence", "Parce que le lancement est effectif, le pipeline de prospects existe et l'entreprise doit desormais standardiser son modele d'execution.", "Page LinkedIn, presse locale, classeur"),
    ("Pourquoi ce modele commercial ?", "Logique de marge", "Les services generent immediatement la tresorerie, tandis que la standardisation des modules d'agents et l'IP creent de la valeur a terme.", "Onglets Offres, P&L, Scenarios"),
    ("Comment evitez-vous le piege des agences de services ?", "Positionnement distinct", "En productisant nos offres, en appliquant des SOPs et en vendant des resultats operationnels au forfait plutot que des heures.", "Onglets Offres, Livraison, SOPs"),
    ("Qui gere reellement l'entreprise ?", "Risque cle-fondateur", "Fenelon dirige la vision et la R&D ; Prince pilote le commercial ; Patricia et Zaina soutiennent la gestion et les operations.", "Onglet Equipe"),
    ("Que faire si le pipeline ralentit ?", "Risque de revenus", "Le modele repose sur plusieurs canaux et prend des conversions tres prudentes. Il ne depend pas d'une source unique.", "Onglet Pipeline"),
    ("Comment les agents IA creent-ils de la valeur ?", "Pertinence technique", "Ils reduisent la friction de notre delivery et constituent nos propres actifs cognitifs (R&D).", "Guide AGENTS, Operations, SOPs"),
    ("Pourquoi le ZAR ?", "Realite du marche", "Parce que nos operations principales se situent dans la zone locale. Les previsions doivent etre realistes et sinceres.", "Hypotheses generales"),
    ("A quoi servira le financement Seed ?", "Utilisation des fonds", "A financer notre force commerciale locale, la capacite technique, la conformite administrative et notre roulement.", "Onglet Financement"),
]
for r, rowv in enumerate(qa_rows, start=5):
    for c, val in enumerate(rowv, start=1):
        ws_qa.cell(row=r, column=c, value=val)
    for c in range(1, 5):
        ws_qa.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
ws_qa.column_dimensions["A"].width = 28
ws_qa.column_dimensions["B"].width = 24
ws_qa.column_dimensions["C"].width = 56
ws_qa.column_dimensions["D"].width = 28


# Dashboard
set_title(ws0, "Tableau de Bord Synthese", "Synthese en une page pour les investisseurs et partenaires : chiffres cles, trajectoire et principes de base.")
dashboard_cards = [
    ("Chiffre d'affaires annee 1", total(base["rev"][:12])),
    ("Marge brute annee 1", total(base["gross"][:12]) / max(total(base["rev"][:12]), 1)),
    ("EBITDA annee 1", total(base["ebitda"][:12])),
    ("Mois d'equilibre (Break-even)", next((i + 1 for i, v in enumerate(base["ebitda"]) if v > 0), inf)),
    ("CA mensuel cible (Mois 36)", base["rev"][-1]),
    ("Tresorerie finale (Mois 36)", base["cash"][-1]),
]
for idx, (label, value) in enumerate(dashboard_cards, start=4):
    ws0.cell(row=idx, column=1, value=label)
    ws0.cell(row=idx, column=2, value=value)
    ws0.cell(row=idx, column=1).font = Font(bold=True)
    if "margin" in label.lower() or "marge" in label.lower():
        fmt_pct(ws0.cell(row=idx, column=2))
    elif "month" in label.lower() or "mois" in label.lower():
        fmt_int(ws0.cell(row=idx, column=2))
    else:
        fmt_money(ws0.cell(row=idx, column=2))

ws0["D4"] = "Synthese executive"
ws0["D5"] = "Forge est notre moteur de cash immediat."
ws0["D6"] = "Sovereign est notre developpement technologique strategique."
ws0["D7"] = "Le modele est conçu pour etre commando (lean) puis scalable."
ws0["D8"] = "Les tarifs sont en ZAR, la TVA est exclue des marges de profit."
ws0["D9"] = "L'impot sur les societes (IS) est modelise a 27%."
ws0["D10"] = "Message cle : Opays Tech est un systeme d'exploitation d'efficience au service des clients."
for r in range(4, 11):
    ws0[f"D{r}"].alignment = Alignment(wrap_text=True)
ws0.column_dimensions["A"].width = 28
ws0.column_dimensions["B"].width = 18
ws0.column_dimensions["D"].width = 42

line = LineChart()
line.title = "Chiffre d'affaires / EBITDA / Tresorerie"
line.y_axis.title = "ZAR"
line.x_axis.title = "Mois"
data = Reference(ws_pnl, min_col=2, max_col=1 + HORIZON, min_row=6, max_row=11)
cats = Reference(ws_pnl, min_col=2, max_col=1 + HORIZON, min_row=4, max_row=4)
line.add_data(data, titles_from_data=False, from_rows=True)
line.set_categories(cats)
line.height = 9
line.width = 18
ws0.add_chart(line, "D12")

bar = BarChart()
bar.title = "Chiffre d'affaires annuel"
bar.y_axis.title = "ZAR"
bar.x_axis.title = "Annee"
data2 = Reference(ws_scn, min_col=2, max_col=4, min_row=5, max_row=7)
cats2 = Reference(ws_scn, min_col=1, max_col=1, min_row=5, max_row=7)
bar.add_data(data2, titles_from_data=False, from_rows=False)
bar.set_categories(cats2)
bar.height = 8
bar.width = 14
ws0.add_chart(bar, "D32")


# formatting
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.freeze_panes or None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                if isinstance(cell.value, str) and len(cell.value) > 60:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")


wb.save(OUT_XLSX)

summary = f"""# Résumé du Modèle Financier Opays Tech

Classeur : `{OUT_XLSX}`

## Chiffres Clés (Scénario Référence / Base)

- Chiffre d'affaires (Année 1) : ZAR {total(base['rev'][:12]):,.0f}
- Marge brute (Année 1) : {total(base['gross'][:12]) / max(total(base['rev'][:12]), 1):.1%}
- EBITDA (Année 1) : ZAR {total(base['ebitda'][:12]):,.0f}
- Mois d'équilibre (Seuil de rentabilité) : {next((i + 1 for i, v in enumerate(base['ebitda']) if v > 0), 'non atteint')}
- CA mensuel cible (Mois 36) : ZAR {base['rev'][-1]:,.0f}
- Trésorerie finale (Mois 36) : ZAR {base['cash'][-1]:,.0f}

## Hypothèses Clés

- Devise : ZAR (Rand sud-africain)
- TVA : 15% traitée en pass-through distinct
- Impôt sur les sociétés (IS) : 27% (hypothèse prudente pour l'Afrique du Sud)
- Taxe sur le chiffre d'affaires : non retenue car les services professionnels en sont exclus selon les règles de la SARS.
- Trésorerie de départ : ZAR {OPENING_CASH:,.0f}
- Levée de fonds recommandée (Seed) : ZAR {SEED_TARGET:,.0f}

## Contenu du Classeur

- Dashboard (Tableau de bord de synthèse)
- Assumptions (Hypothèses opérationnelles modifiables)
- Offers (Détail commercial des offres)
- GTM (Canaux de vente et taux de conversion)
- Revenue_Model (Modélisation mensuelle des revenus)
- Expenses (Détail des charges fixes par phase)
- P&L (Compte de résultat prévisionnel)
- Cashflow (Flux de trésorerie mensuels)
- Funding (Détail de l'utilisation des fonds)
- Scenarios (Comparaison conservateur / base / optimiste)
- Risks_QA (Registre des risques et Q&A)
- Sources (Références fiscales et réglementaires)
- Operations (Principes opérationnels et rôles)
- Capacity (Modélisation de la charge de travail)
- Delivery (Déroulement du parcours client)
- Cadence (Rythme hebdomadaire d'exécution)
- SOPs (Procédures opérationnelles standardisées)
- Team (Profil de l'équipe et répartition)
- Pipeline (Funnel de conversion)
- 90_Day_Plan (Plan d'action pour le premier trimestre)
- Pitch_Outline (Structure du deck de présentation)
- Investor_QA (Réponses aux questions sensibles des investisseurs)

## Couche Opérationnelle

Le classeur ne se limite pas à des chiffres. Il définit le système d'exploitation d'Opays Tech :

- les principes de travail et la répartition des rôles,
- la modélisation de la capacité réelle de l'équipe pour rester réaliste,
- le playbook de livraison du lead jusqu'à l'abonnement récurrent,
- le rythme hebdomadaire pour traduire la stratégie en routine,
- les procédures standard (SOPs) pour industrialiser la livraison de valeur,
- la structure de l'équipe alignée sur la charge de travail réelle,
- le plan d'action des 90 premiers jours pour ancrer l'exécution,
- le registre des risques et les réponses argumentées aux investisseurs.

## Narratif de Présentation aux Investisseurs

Opays Tech est modélisé comme un cabinet d'ingénierie de l'efficience sobre et rentable :

1. La branche **Forge** génère la trésorerie immédiate via des audits, des sprints et des retainers.
2. La branche **Sovereign** valorise l'expertise via des dossiers de recherche et R&D premium.
3. Le modèle reste très prudent sur les charges fixes et applique strictement la conformité fiscale locale.
4. L'entreprise évite le "piège des agences classiques" en vendant des résultats productisés et en cadrant strictement son périmètre.

"""
OUT_MD.write_text(summary, encoding="utf-8")

print(f"Saved workbook to {OUT_XLSX}")
print(f"Saved summary to {OUT_MD}")
